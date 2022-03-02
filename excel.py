from openpyxl import load_workbook

# 获取表
file_location = 'C:/Users/shmily/Desktop/data.xlsx'
excel = load_workbook(file_location)
table = excel["Sheet1"]

# 获取行数、列数
rows = table.max_row
columns = table.max_column
